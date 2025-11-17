import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
import datetime
import os

# --- Rutas de Archivos Estáticos (Logo, Firma) ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(BASE_DIR, "plantillas", "logo_liderman.png")
FIRMAS_DIR = os.path.join(BASE_DIR, "plantillas", "firmas")

# Dimensiones originales de las firmas e imágenes (ancho x alto en píxeles)
FIRMAS_DIMENSIONES = {
    'logo_liderman.png': (333, 150),  # Logo principal
    'firma_Brenda_Bendezu.png': (64, 37),
    'firma_Daniel_Perez.png': (106, 48),
    'firma_eliana.png': (250, 100),
    'firma_Fredy_Gutierrez.png': (97, 52),
    'firma_Gabriela_Garcia.png': (119, 64),
    'firma_Indira_Merino.png': (57, 56),
    'firma_Jose_Alvines.png': (82, 43),
    'firma_Juan_Terrones.png': (52, 37),
    'firma_Luis_Sanchez.png': (87, 62),
    'firma_Renzo_Asian.png': (106, 49),
    'firma_Ruth_Ponce.png': (118, 67),
    'firma_Segundo_Maldonado.png': (61, 36),
    'firma_Yesenia_Armacanqui.png': (61, 33),
    'firma_capacitador.png': (82, 43),  # Default (mismo que Jose Alvines)
}

def get_firma_path(nombre_archivo):
    """
    Obtiene la ruta completa de un archivo de firma.
    Si no existe, busca la firma por defecto.
    """
    # Intentar con el nombre específico
    firma_path = os.path.join(FIRMAS_DIR, nombre_archivo)
    if os.path.exists(firma_path):
        return firma_path
    
    # Si no existe, buscar en el directorio de plantillas
    firma_path = os.path.join(BASE_DIR, "plantillas", nombre_archivo)
    if os.path.exists(firma_path):
        return firma_path
    
    # Firma por defecto
    default_path = os.path.join(BASE_DIR, "plantillas", "firma_capacitador.png")
    if os.path.exists(default_path):
        return default_path
    
    return None

def apply_border_to_range(ws, start_cell, end_cell, border):
    """
    Aplica bordes a todas las celdas en un rango, incluso si están combinadas.
    """
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")
    
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).border = border

def calculate_optimal_image_size(original_width, original_height, max_width, max_height, fill_percent=0.8, allow_compress=False):
    """
    Calcula el tamaño óptimo de una imagen para que ocupe el fill_percent del espacio disponible.
    
    Args:
        original_width: Ancho original de la imagen en píxeles
        original_height: Alto original de la imagen en píxeles
        max_width: Ancho máximo disponible en píxeles (del espacio de la celda)
        max_height: Alto máximo disponible en píxeles (del espacio de la celda)
        fill_percent: Porcentaje del espacio a ocupar (default 0.8 = 80%)
        allow_compress: Si True, permite comprimir la imagen si es necesario. Si False, solo estira.
    
    Returns:
        tuple: (width, height) en píxeles
    """
    # Calcular espacio objetivo (80% del disponible)
    target_width = max_width * fill_percent
    target_height = max_height * fill_percent
    
    # Calcular ratio de aspecto
    aspect_ratio = original_width / original_height
    
    # Escalar manteniendo proporción
    if target_width / aspect_ratio <= target_height:
        # Limitado por ancho
        final_width = target_width
        final_height = target_width / aspect_ratio
    else:
        # Limitado por alto
        final_height = target_height
        final_width = target_height * aspect_ratio
    
    # Si no se permite comprimir y la imagen es más grande, usar tamaño original
    if not allow_compress:
        if final_width < original_width or final_height < original_height:
            # La imagen es más pequeña que el espacio, estirar
            final_width = max(final_width, original_width)
            final_height = max(final_height, original_height)
            # Reajustar para mantener proporción si se pasó
            if final_width > target_width or final_height > target_height:
                if final_width / aspect_ratio <= target_height:
                    final_height = final_width / aspect_ratio
                else:
                    final_width = final_height * aspect_ratio
    
    return int(final_width), int(final_height)

def add_centered_image(ws, img_path, cell_address, img_width, img_height):
    """
    Agrega una imagen con el tamaño especificado en una celda.
    Usa método simple y confiable para asegurar que la imagen aparezca correctamente.
    """
    try:
        img = OpenpyxlImage(img_path)
        # Establecer tamaño exacto de la imagen
        img.width = img_width
        img.height = img_height
        # Agregar imagen en la celda especificada
        ws.add_image(img, cell_address)
        return True
    except Exception as e:
        print(f"Error: No se pudo agregar imagen: {e}")
        return False

def add_truly_centered_image(ws, img_path, start_col, start_row, end_col, end_row, img_width, img_height, cell_width_px, cell_height_px):
    """
    Agrega una imagen verdaderamente centrada en un rango de celdas usando TwoCellAnchor.
    
    Args:
        ws: worksheet
        img_path: ruta de la imagen
        start_col: columna inicial (0-indexed)
        start_row: fila inicial (0-indexed)
        end_col: columna final (0-indexed)
        end_row: fila final (0-indexed)
        img_width: ancho de la imagen en píxeles
        img_height: alto de la imagen en píxeles
        cell_width_px: ancho total del rango de celdas en píxeles
        cell_height_px: alto total del rango de celdas en píxeles
    """
    try:
        from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
        from openpyxl.utils.units import pixels_to_EMU
        
        # Calcular offsets para centrar (asegurar que sean positivos)
        offset_x = max(0, (cell_width_px - img_width) / 2)
        offset_y = max(0, (cell_height_px - img_height) / 2)
        
        # Crear imagen
        img = OpenpyxlImage(img_path)
        img.width = img_width
        img.height = img_height
        
        # Crear anclaje centrado
        _from = AnchorMarker(col=start_col, row=start_row, colOff=pixels_to_EMU(offset_x), rowOff=pixels_to_EMU(offset_y))
        to = AnchorMarker(col=start_col, row=start_row, colOff=pixels_to_EMU(offset_x + img_width), rowOff=pixels_to_EMU(offset_y + img_height))
        img.anchor = TwoCellAnchor(editAs='oneCell', _from=_from, to=to)
        
        ws.add_image(img)
        return True
    except Exception as e:
        print(f"Error: No se pudo agregar imagen centrada: {e}")
        import traceback
        traceback.print_exc()
        return False

def add_logo_liderman(ws):
    """
    Agrega el logo Liderman en A1:B4 con tamaño grande y visible.
    """
    try:
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.utils.units import pixels_to_EMU
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        
        if not os.path.exists(LOGO_PATH):
            print(f"ADVERTENCIA: Logo no encontrado en {LOGO_PATH}")
            return False
        
        # Dimensiones del logo ajustadas para caber en A1:B4
        logo_width = 290
        logo_height = 70
        
        # Posición con offset para centrar visualmente en A1:B4
        offset_x = 95  # Mover a la derecha para centrar
        offset_y = 3   # Mover hacia abajo para centrar
        
        # Crear imagen
        img = OpenpyxlImage(LOGO_PATH)
        print(f"DEBUG: Logo con offset_x={offset_x}, offset_y={offset_y}, tamaño={logo_width}x{logo_height}")
        
        # Usar OneCellAnchor con tamaño fijo (NO se comprime)
        marker = AnchorMarker(col=0, row=0, colOff=pixels_to_EMU(offset_x), rowOff=pixels_to_EMU(offset_y))
        size = XDRPositiveSize2D(pixels_to_EMU(logo_width), pixels_to_EMU(logo_height))
        img.anchor = OneCellAnchor(_from=marker, ext=size)
        
        ws.add_image(img)
        print(f"✓ Logo agregado: {logo_width}x{logo_height}px en posición ({offset_x}, {offset_y})")
        return True
        
    except Exception as e:
        print(f"ERROR LOGO: {e}")
        import traceback
        traceback.print_exc()
        return False

def add_firma_eliana(ws, footer_row):
    """
    Agrega la firma de Eliana en G:I del footer con tamaño grande y visible.
    """
    try:
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.utils.units import pixels_to_EMU
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        
        firma_path = get_firma_path('firma_eliana.png')
        if not firma_path:
            print("ADVERTENCIA: Firma Eliana no encontrada")
            return False
        
        # Dimensiones de la firma (80% del tamaño original 285x132)
        firma_width = 228
        firma_height = 106
        
        # Posición con offset para centrar en G:I (3 columnas)
        offset_x = 48   # Centrar en el espacio G:I
        offset_y = 14   # Centrar verticalmente
        
        # Crear imagen
        img = OpenpyxlImage(firma_path)
        
        # Usar OneCellAnchor con tamaño fijo (NO se comprime)
        marker = AnchorMarker(col=6, row=footer_row-1, colOff=pixels_to_EMU(offset_x), rowOff=pixels_to_EMU(offset_y))
        size = XDRPositiveSize2D(pixels_to_EMU(firma_width), pixels_to_EMU(firma_height))
        img.anchor = OneCellAnchor(_from=marker, ext=size)
        
        ws.add_image(img)
        print(f"✓ Firma agregada: {firma_width}x{firma_height}px en posición ({offset_x}, {offset_y})")
        return True
        
    except Exception as e:
        print(f"ERROR FIRMA: {e}")
        import traceback
        traceback.print_exc()
        return False

def create_formatted_excel(df_course, course_details):
    """
    Toma un DataFrame y detalles de curso, y devuelve 
    un archivo Excel (en bytes) con el formato exacto.
    """
    try:
        output = BytesIO()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = course_details['Nombre Curso'][:31]

        # --- Definir Estilos ---
        font_bold_14 = Font(bold=True, size=14)
        font_bold_10 = Font(bold=True, size=10)
        font_normal = Font(size=10)

        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        top_center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
        center_left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)  # Centrado vertical, izquierda horizontal

        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # --- Ajustar Anchos de Columna ---
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 52
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 57
        ws.column_dimensions['E'].width = 19
        ws.column_dimensions['F'].width = 19
        ws.column_dimensions['G'].width = 19
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12 

        # --- Fila 1-4: Logo y Encabezado ---
        ws.merge_cells('A1:B4')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        apply_border_to_range(ws, 'A1', 'B4', thin_border)  # Agregar bordes al logo
        
        # Ajustar altura de las filas del logo para acomodar logo grande
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 22
        ws.row_dimensions[4].height = 22
        
        # Agregar logo Liderman usando función específica
        add_logo_liderman(ws)

        ws.merge_cells('C1:G4')
        cell = ws['C1']
        cell.value = "FORMATO\n\nLISTA DE ASISTENCIA VIRTUAL"
        cell.font = font_bold_14
        cell.alignment = center_align
        apply_border_to_range(ws, 'C1', 'G4', thin_border)

        ws.cell(row=1, column=8, value="Código:").font = font_bold_10
        ws.cell(row=1, column=8).border = thin_border
        ws.cell(row=1, column=9, value="JV-GTH-F-111").font = font_normal
        ws.cell(row=1, column=9).border = thin_border

        ws.cell(row=2, column=8, value="Versión:").font = font_bold_10
        ws.cell(row=2, column=8).border = thin_border
        ws.cell(row=2, column=9, value=4).font = font_normal
        ws.cell(row=2, column=9).alignment = center_align  # Centrar el valor
        ws.cell(row=2, column=9).border = thin_border

        ws.cell(row=3, column=8, value="Fecha:").font = font_bold_10
        ws.cell(row=3, column=8).border = thin_border
        ws.cell(row=3, column=9, value="27/12/2022").font = font_normal
        ws.cell(row=3, column=9).border = thin_border

        ws.cell(row=4, column=8, value="Página 01").font = font_normal
        ws.cell(row=4, column=8).border = thin_border
        ws.cell(row=4, column=9, value="").border = thin_border

        # --- Fila 5: Datos del empleador ---
        ws.merge_cells('A5:I5')
        cell = ws['A5']
        cell.value = "Datos del empleador:"
        cell.font = font_bold_10
        cell.alignment = left_align
        apply_border_to_range(ws, 'A5', 'I5', thin_border)

        # --- FILA 6: Cabeceras de la tabla superior ---
        ws.cell(row=6, column=1, value="MARCAR").font = font_bold_10
        ws.cell(row=6, column=1).alignment = center_align
        ws.cell(row=6, column=1).border = thin_border
        
        ws.cell(row=6, column=2, value="RAZÓN SOCIAL").font = font_bold_10
        ws.cell(row=6, column=2).alignment = center_align
        ws.cell(row=6, column=2).border = thin_border
        
        ws.cell(row=6, column=3, value="RUC").font = font_bold_10
        ws.cell(row=6, column=3).alignment = center_align
        ws.cell(row=6, column=3).border = thin_border
        
        ws.merge_cells('D6:G6')
        ws.cell(row=6, column=4, value="DOMICILIO").font = font_bold_10
        ws.cell(row=6, column=4).alignment = center_align
        apply_border_to_range(ws, 'D6', 'G6', thin_border)
        
        ws.merge_cells('H6:I6')
        ws.cell(row=6, column=8, value="ACTIVIDAD ECONOMICA").font = font_bold_10
        ws.cell(row=6, column=8).alignment = center_align
        apply_border_to_range(ws, 'H6', 'I6', thin_border)

        # --- FILAS 7-10: Datos de las empresas ---
        # Fila 7
        ws.cell(row=7, column=1, value="X").font = font_normal
        ws.cell(row=7, column=1).alignment = center_align
        ws.cell(row=7, column=1).border = thin_border
        
        ws.cell(row=7, column=2, value="J&V RESGUARDO SAC").font = font_normal
        ws.cell(row=7, column=2).alignment = center_align
        ws.cell(row=7, column=2).border = thin_border
        
        ws.cell(row=7, column=3, value="20100901481").font = font_normal
        ws.cell(row=7, column=3).alignment = center_align
        ws.cell(row=7, column=3).border = thin_border
        
        ws.merge_cells('D7:G7')
        ws.cell(row=7, column=4, value="AV. DEFENSORES DEL MORRO N°1620 - CHORRILLOS").font = font_normal
        ws.cell(row=7, column=4).alignment = center_align
        apply_border_to_range(ws, 'D7', 'G7', thin_border)
        
        ws.merge_cells('H7:I7')
        ws.cell(row=7, column=8, value="VIGILANCIA PRIVADA").font = font_normal
        ws.cell(row=7, column=8).alignment = center_align
        apply_border_to_range(ws, 'H7', 'I7', thin_border)
        
        # Fila 8
        ws.cell(row=8, column=1).border = thin_border
        ws.cell(row=8, column=2, value="J&V RESGUARDO SELVA SAC").font = font_normal
        ws.cell(row=8, column=2).alignment = center_align
        ws.cell(row=8, column=2).border = thin_border
        
        ws.cell(row=8, column=3, value="20493762789").font = font_normal
        ws.cell(row=8, column=3).alignment = center_align
        ws.cell(row=8, column=3).border = thin_border
        
        ws.merge_cells('D8:G8')
        ws.cell(row=8, column=4, value="JR. NAUTA 269 - IQUITOS").font = font_normal
        ws.cell(row=8, column=4).alignment = center_align
        apply_border_to_range(ws, 'D8', 'G8', thin_border)
        
        ws.merge_cells('H8:I8')
        ws.cell(row=8, column=8, value="VIGILANCIA PRIVADA").font = font_normal
        ws.cell(row=8, column=8).alignment = center_align
        apply_border_to_range(ws, 'H8', 'I8', thin_border)

        # Fila 9
        ws.cell(row=9, column=1).border = thin_border
        ws.cell(row=9, column=2, value="LIDERMAN SERVICIOS SAC").font = font_normal
        ws.cell(row=9, column=2).alignment = center_align
        ws.cell(row=9, column=2).border = thin_border
        
        ws.cell(row=9, column=3, value="20601355761").font = font_normal
        ws.cell(row=9, column=3).alignment = center_align
        ws.cell(row=9, column=3).border = thin_border
        
        ws.merge_cells('D9:G9')
        ws.cell(row=9, column=4, value="AV. DEFENSORES DEL MORRO N°1620 - CHORRILLOS").font = font_normal
        ws.cell(row=9, column=4).alignment = center_align
        apply_border_to_range(ws, 'D9', 'G9', thin_border)
        
        ws.merge_cells('H9:I9')
        ws.cell(row=9, column=8, value="ACTIVIDADES DE TRANSPORTE").font = font_normal
        ws.cell(row=9, column=8).alignment = center_align
        apply_border_to_range(ws, 'H9', 'I9', thin_border)
        
        # Fila 10
        ws.cell(row=10, column=1).border = thin_border
        ws.cell(row=10, column=2, value="J&V ALARMAS S.A.C.").font = font_normal
        ws.cell(row=10, column=2).alignment = center_align
        ws.cell(row=10, column=2).border = thin_border
        
        ws.cell(row=10, column=3, value="20303166573").font = font_normal
        ws.cell(row=10, column=3).alignment = center_align
        ws.cell(row=10, column=3).border = thin_border
        
        ws.merge_cells('D10:G10')
        ws.cell(row=10, column=4, value="AV. DEFENSORES DEL MORRO N°1620 - CHORRILLOS").font = font_normal
        ws.cell(row=10, column=4).alignment = center_align
        apply_border_to_range(ws, 'D10', 'G10', thin_border)
        
        ws.merge_cells('H10:I10')
        ws.cell(row=10, column=8, value="ACTIVIDAD DE INVESTIGACIÓN").font = font_normal
        ws.cell(row=10, column=8).alignment = center_align
        apply_border_to_range(ws, 'H10', 'I10', thin_border)

        # --- FILA 11: Tema/Motivo ---
        ws.row_dimensions[11].height = 39.60
        
        ws.merge_cells('A11:B11')
        ws['A11'].value = "Tema/Motivo:"
        ws['A11'].font = font_bold_10
        ws['A11'].alignment = center_align
        apply_border_to_range(ws, 'A11', 'B11', thin_border)
        
        ws.merge_cells('C11:E11')
        ws['C11'].value = course_details['Tema/Motivo']
        ws['C11'].font = font_bold_10  # NEGRITA para el nombre del tema/motivo
        ws['C11'].alignment = center_align
        apply_border_to_range(ws, 'C11', 'E11', thin_border)
        
        ws.merge_cells('F11:G11')
        ws['F11'].value = "Grabación/ Material:"
        ws['F11'].font = font_bold_10
        ws['F11'].alignment = center_align
        apply_border_to_range(ws, 'F11', 'G11', thin_border)
        
        ws.merge_cells('H11:I11')
        ws['H11'].value = course_details['Grabacion/ Material']
        ws['H11'].font = font_normal
        ws['H11'].alignment = center_align
        apply_border_to_range(ws, 'H11', 'I11', thin_border)

        # --- FILA 12: Contenido ---
        # Calcular altura automática basada en el contenido
        contenido = course_details['Contenido/ Sub Temas']
        # Ancho total de las columnas C a I en caracteres (aprox)
        ancho_disponible = 20 + 57 + 19 + 19 + 19 + 12 + 12  # suma de anchos C-I
        # Calcular número de líneas necesarias
        num_lineas = contenido.count('\n') + 1  # contar saltos de línea
        # Estimar líneas adicionales por wrap de texto
        chars_por_linea = ancho_disponible * 1.2  # ajuste por tamaño de fuente
        lineas_por_wrap = len(contenido) / chars_por_linea
        total_lineas = max(num_lineas, lineas_por_wrap)
        # Altura de fila: ~15 puntos por línea + padding
        altura_calculada = max(50, min(400, total_lineas * 15 + 10))  # mínimo 50, máximo 400
        ws.row_dimensions[12].height = altura_calculada
        
        ws.merge_cells('A12:B12')
        ws['A12'].value = "Contenido/ Sub Temas:"
        ws['A12'].font = font_bold_10
        ws['A12'].alignment = center_align  # Centrado completamente para el encabezado
        apply_border_to_range(ws, 'A12', 'B12', thin_border)
        
        ws.merge_cells('C12:I12')
        ws['C12'].value = contenido
        ws['C12'].font = font_normal
        ws['C12'].alignment = center_left_align  # Centrado vertical, alineado a la izquierda
        apply_border_to_range(ws, 'C12', 'I12', thin_border)

        # --- FILA 13: Capacitador/Firma ---
        ws.row_dimensions[13].height = 95  # Aumentar altura para firmas más grandes
        
        ws.merge_cells('A13:B13')
        ws['A13'].value = "Capacitador/Entrenador:"
        ws['A13'].font = font_bold_10
        ws['A13'].alignment = center_align
        apply_border_to_range(ws, 'A13', 'B13', thin_border)
        
        ws.merge_cells('C13:D13')
        ws['C13'].value = course_details['Capacitador/Entrenador']
        ws['C13'].font = font_normal
        ws['C13'].alignment = center_align
        apply_border_to_range(ws, 'C13', 'D13', thin_border)
        
        ws['E13'].value = "Firma:"
        ws['E13'].font = font_bold_10
        ws['E13'].alignment = center_align
        ws['E13'].border = thin_border
        
        # Firma (Imagen) - Manejar una o múltiples firmas
        ws['F13'].border = thin_border
        ws['F13'].alignment = Alignment(horizontal='center', vertical='center')
        
        firmas_str = course_details.get('Firma', 'firma_capacitador.png')
        # Separar firmas si hay múltiples (separadas por |)
        firmas_list = [f.strip() for f in firmas_str.split('|')]
        
        # Espacio disponible en F13: 143x127 px (95 puntos = 127px) → usar 85% para firmas más grandes sin comprimir
        if len(firmas_list) == 1:
            # Una sola firma - centrada en F13
            firma_path = get_firma_path(firmas_list[0])
            if firma_path:
                # Obtener dimensiones originales
                firma_nombre = os.path.basename(firma_path)
                orig_w, orig_h = FIRMAS_DIMENSIONES.get(firma_nombre, (82, 43))
                # Calcular tamaño óptimo para F13 (NO comprimir, solo estirar si es necesario)
                firma_w, firma_h = calculate_optimal_image_size(orig_w, orig_h, 143, 127, 0.85, allow_compress=False)
                # F13 = columna 5, fila 12 (0-indexed), 143x127 px
                add_truly_centered_image(ws, firma_path, 5, 12, 5, 12, firma_w, firma_h, 143, 127)
        elif len(firmas_list) == 2:
            # Dos firmas - lado a lado en F13 (cada una ocupa 50% del ancho)
            firma1_path = get_firma_path(firmas_list[0])
            firma2_path = get_firma_path(firmas_list[1])
            
            if firma1_path and firma2_path:
                from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
                from openpyxl.utils.units import pixels_to_EMU
                
                # Espacio para cada firma: 70x127 px (mitad del ancho, altura completa 95pts = 127px) → NO comprimir
                firma1_nombre = os.path.basename(firma1_path)
                firma2_nombre = os.path.basename(firma2_path)
                orig_w1, orig_h1 = FIRMAS_DIMENSIONES.get(firma1_nombre, (82, 43))
                orig_w2, orig_h2 = FIRMAS_DIMENSIONES.get(firma2_nombre, (82, 43))
                
                # Calcular tamaños óptimos MÁS GRANDES (90% del espacio para mayor visibilidad, sin comprimir)
                firma1_w, firma1_h = calculate_optimal_image_size(orig_w1, orig_h1, 70, 127, 0.90, allow_compress=False)
                firma2_w, firma2_h = calculate_optimal_image_size(orig_w2, orig_h2, 70, 127, 0.90, allow_compress=False)
                
                # Centrar las dos firmas horizontalmente en F13 (143px de ancho total)
                # Calcular espaciado para centrado: (143 - firma1_w - firma2_w) / 3 para dejar espacios iguales
                espacio_horizontal = max(5, (143 - firma1_w - firma2_w) / 3)
                offset_vertical = (127 - max(firma1_h, firma2_h)) / 2  # Centrar verticalmente
                
                # Primera firma a la izquierda con margen
                img1 = OpenpyxlImage(firma1_path)
                img1.width = firma1_w
                img1.height = firma1_h
                _from1 = AnchorMarker(col=5, row=12, colOff=pixels_to_EMU(espacio_horizontal), rowOff=pixels_to_EMU(offset_vertical))
                to1 = AnchorMarker(col=5, row=12, colOff=pixels_to_EMU(espacio_horizontal + firma1_w), rowOff=pixels_to_EMU(offset_vertical + firma1_h))
                img1.anchor = TwoCellAnchor(editAs='oneCell', _from=_from1, to=to1)
                ws.add_image(img1)
                
                # Segunda firma a la derecha con margen
                img2 = OpenpyxlImage(firma2_path)
                img2.width = firma2_w
                img2.height = firma2_h
                offset_horizontal_2 = espacio_horizontal * 2 + firma1_w
                _from2 = AnchorMarker(col=5, row=12, colOff=pixels_to_EMU(offset_horizontal_2), rowOff=pixels_to_EMU(offset_vertical))
                to2 = AnchorMarker(col=5, row=12, colOff=pixels_to_EMU(offset_horizontal_2 + firma2_w), rowOff=pixels_to_EMU(offset_vertical + firma2_h))
                img2.anchor = TwoCellAnchor(editAs='oneCell', _from=_from2, to=to2)
                ws.add_image(img2)
            elif firma1_path:
                # Solo primera firma disponible
                firma_nombre = os.path.basename(firma1_path)
                orig_w, orig_h = FIRMAS_DIMENSIONES.get(firma_nombre, (82, 43))
                firma_w, firma_h = calculate_optimal_image_size(orig_w, orig_h, 143, 127, 0.85, allow_compress=False)
                add_truly_centered_image(ws, firma1_path, 5, 12, 5, 12, firma_w, firma_h, 143, 127)
            elif firma2_path:
                # Solo segunda firma disponible
                firma_nombre = os.path.basename(firma2_path)
                orig_w, orig_h = FIRMAS_DIMENSIONES.get(firma_nombre, (82, 43))
                firma_w, firma_h = calculate_optimal_image_size(orig_w, orig_h, 143, 127, 0.85, allow_compress=False)
                add_truly_centered_image(ws, firma2_path, 5, 12, 5, 12, firma_w, firma_h, 143, 127)
        
        ws['G13'].value = "Duración:"
        ws['G13'].font = font_bold_10
        ws['G13'].alignment = center_align
        ws['G13'].border = thin_border
        
        ws.merge_cells('H13:I13')
        ws['H13'].value = course_details['Duracion']
        ws['H13'].font = font_normal
        ws['H13'].alignment = center_align
        apply_border_to_range(ws, 'H13', 'I13', thin_border)

        # --- FILA 14: Motivo ---
        ws.merge_cells('A14:I14')
        ws['A14'].value = "Motivo:"
        ws['A14'].font = font_bold_10
        ws['A14'].alignment = left_align
        apply_border_to_range(ws, 'A14', 'I14', thin_border)

        # --- FILA 15: Tipo de actividad y N° de Trabajadores ---
        ws.merge_cells('A15:F15')
        ws['A15'].value = "Inducción ( ) Capacitación (X) Entrenamiento ( ) Charla de 5 minutos ( )"
        ws['A15'].font = font_normal
        ws['A15'].alignment = left_align
        apply_border_to_range(ws, 'A15', 'F15', thin_border)
        
        ws['G15'].value = "N° de Trabajadores:"
        ws['G15'].font = font_bold_10
        ws['G15'].alignment = center_align
        ws['G15'].border = thin_border
        
        ws.merge_cells('H15:I15')
        ws['H15'].value = len(df_course)
        ws['H15'].font = font_normal
        ws['H15'].alignment = center_align
        apply_border_to_range(ws, 'H15', 'I15', thin_border)

        # --- FILA 16: Cabeceras de la tabla de datos ---
        data_header_row = 16
        headers = ["N°", "Apellidos y Nombres", "DNI", "Unidad (Cliente)", "Nota", "Fecha Examen", "Hora Conexión"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=data_header_row, column=col_idx, value=header)
            cell.font = font_bold_10
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = grey_fill
        
        # Agregar columna "Observación" que abarca H16:I16
        ws.merge_cells('H16:I16')
        ws['H16'].value = "Observación"
        ws['H16'].font = font_bold_10
        ws['H16'].alignment = center_align
        ws['H16'].fill = grey_fill
        apply_border_to_range(ws, 'H16', 'I16', thin_border)

        # --- Pegar los datos del DataFrame ---
        rows = df_course.values.tolist()
        for row_idx, row_data in enumerate(rows, start=data_header_row + 1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.font = font_normal
                # Todos los datos centrados (incluyendo nombres en columna 2)
                cell.alignment = center_align
                cell.border = thin_border
            
            # Agregar celdas para la columna "Observación" (H e I) con bordes
            ws.cell(row=row_idx, column=8).border = thin_border
            ws.cell(row=row_idx, column=9).border = thin_border
            # Combinar las celdas H e I para Observación en cada fila de datos
            ws.merge_cells(f'H{row_idx}:I{row_idx}')
            ws.cell(row=row_idx, column=8).alignment = center_align

        # --- Pie de página: Responsable del Registro ---
        # Comenzar exactamente después de la última fila de datos (sin espacios)
        footer_start_row = data_header_row + len(df_course) + 1
        
        # FILA ENCABEZADO: "Responsable del Registro:" con fondo gris
        ws.merge_cells(f'A{footer_start_row}:I{footer_start_row}')
        ws[f'A{footer_start_row}'].value = "Responsable del Registro:"
        ws[f'A{footer_start_row}'].font = font_bold_10
        ws[f'A{footer_start_row}'].alignment = left_align
        ws[f'A{footer_start_row}'].fill = grey_fill
        apply_border_to_range(ws, f'A{footer_start_row}', f'I{footer_start_row}', thin_border)
        
        # FILA 1 del pie: Apellidos y Nombres + Firma + Imagen (ahora en footer_start_row + 1)
        ws.merge_cells(f'A{footer_start_row+1}:E{footer_start_row+1}')
        # Usar Rich Text para combinar negrita y normal
        from openpyxl.cell.text import InlineFont
        from openpyxl.cell.rich_text import TextBlock, CellRichText
        
        # Crear texto con formato mixto: "Apellidos y Nombres:" en negrita, resto normal
        rich_text = CellRichText()
        rich_text.append(TextBlock(InlineFont(b=True, sz=10), "Apellidos y Nombres: "))
        rich_text.append(TextBlock(InlineFont(sz=10), "Cuaila Colque, Eliana"))
        ws[f'A{footer_start_row+1}'].value = rich_text
        ws[f'A{footer_start_row+1}'].alignment = center_left_align  # Centrado vertical, pegado a izquierda
        apply_border_to_range(ws, f'A{footer_start_row+1}', f'E{footer_start_row+1}', thin_border)
        
        ws[f'F{footer_start_row+1}'].value = "Firma:"
        ws[f'F{footer_start_row+1}'].font = font_bold_10
        ws[f'F{footer_start_row+1}'].alignment = center_align
        ws[f'F{footer_start_row+1}'].border = thin_border
        
        # Imagen de firma en G:I (centrada y más grande)
        ws.merge_cells(f'G{footer_start_row+1}:I{footer_start_row+1}')
        ws[f'G{footer_start_row+1}'].alignment = center_align
        apply_border_to_range(ws, f'G{footer_start_row+1}', f'I{footer_start_row+1}', thin_border)
        
        # Ajustar altura de fila para la firma Eliana
        ws.row_dimensions[footer_start_row+1].height = 100
        
        # Agregar firma Eliana (FIRMA PRINCIPAL) usando función específica (ahora en footer_start_row+1)
        add_firma_eliana(ws, footer_start_row+1)
        
        # FILA 2 del pie: Cargo + Fecha (ahora en footer_start_row+2)
        ws.merge_cells(f'A{footer_start_row+2}:E{footer_start_row+2}')
        # Crear texto con formato mixto: "Cargo:" en negrita, resto normal
        rich_text_cargo = CellRichText()
        rich_text_cargo.append(TextBlock(InlineFont(b=True, sz=10), "Cargo: "))
        rich_text_cargo.append(TextBlock(InlineFont(sz=10), "Coordinadora de Capacitación y Desarrollo"))
        ws[f'A{footer_start_row+2}'].value = rich_text_cargo
        ws[f'A{footer_start_row+2}'].alignment = center_left_align  # Centrado vertical, pegado a izquierda
        apply_border_to_range(ws, f'A{footer_start_row+2}', f'E{footer_start_row+2}', thin_border)
        
        ws[f'F{footer_start_row+2}'].value = "Fecha:"
        ws[f'F{footer_start_row+2}'].font = font_bold_10
        ws[f'F{footer_start_row+2}'].alignment = center_align
        ws[f'F{footer_start_row+2}'].border = thin_border
        
        ws.merge_cells(f'G{footer_start_row+2}:I{footer_start_row+2}')
        ws[f'G{footer_start_row+2}'].value = datetime.date.today().strftime('%d/%m/%Y')
        ws[f'G{footer_start_row+2}'].font = font_normal
        ws[f'G{footer_start_row+2}'].alignment = center_align
        apply_border_to_range(ws, f'G{footer_start_row+2}', f'I{footer_start_row+2}', thin_border)

        # --- CONFIGURACIÓN DE PÁGINA PARA IMPRESIÓN/PDF ---
        # Configurar para A4: columnas en 1 página, filas en múltiples páginas
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # Orientación vertical
        ws.page_setup.paperSize = ws.PAPERSIZE_A4  # Tamaño A4
        ws.page_setup.fitToPage = True  # Ajustar a página (solo ancho)
        ws.page_setup.fitToHeight = 0  # 0 = Sin límite de páginas verticales
        ws.page_setup.fitToWidth = 1   # 1 = Todas las columnas en 1 página de ancho
        
        # Configurar márgenes (en pulgadas) - más estrechos para aprovechar espacio
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.25
        ws.page_margins.bottom = 0.25
        ws.page_margins.header = 0.1
        ws.page_margins.footer = 0.1
        
        # Configurar área de impresión (desde A1 hasta la última celda con datos)
        last_row = footer_start_row + 2  # Ahora son 3 filas de footer (encabezado + 2 filas de datos)
        ws.print_area = f'A1:I{last_row}'
        
        # NO centrar en la página - alinear al inicio (arriba-izquierda)
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = False  # No centrar horizontalmente
        ws.print_options.verticalCentered = False    # No centrar verticalmente

        # Guardar el libro de trabajo en el buffer de memoria
        wb.save(output)
        return output.getvalue()

    except Exception as e:
        print(f"Error fatal al crear Excel: {e}")
        return None